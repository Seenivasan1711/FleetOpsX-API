import csv
import io
from datetime import datetime, time
from typing import Any
from uuid import UUID
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.driver import Driver
from app.models.vehicle import Vehicle
from app.models.depot import Depot
from app.models.order import Order


_REQUIRED = {"delivery_address", "scheduled_date"}
_PRIORITY_VALUES = {"LOW", "NORMAL", "HIGH", "CRITICAL"}


def _parse_time(val: Any) -> time | None:
    if not val:
        return None
    if isinstance(val, time):
        return val
    s = str(val).strip()
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            pass
    return None


def _parse_float(val: Any) -> float | None:
    if val is None or str(val).strip() == "":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _parse_bool(val: Any) -> bool:
    if isinstance(val, bool):
        return val
    return str(val).strip().lower() in ("yes", "true", "1")


def import_orders_from_excel(db: Session, tenant_id: UUID, file_bytes: bytes) -> dict:
    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"created": 0, "errors": ["Empty file"]}

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    created = 0
    errors: list[str] = []

    for row_idx, row in enumerate(rows[1:], start=2):
        row_data = dict(zip(headers, row))

        # validate required
        missing = [f for f in _REQUIRED if not row_data.get(f)]
        if missing:
            errors.append(f"Row {row_idx}: missing required fields: {', '.join(missing)}")
            continue

        # parse scheduled_date
        raw_date = row_data.get("scheduled_date")
        if isinstance(raw_date, datetime):
            scheduled_date = raw_date
        else:
            try:
                scheduled_date = datetime.strptime(str(raw_date).strip(), "%Y-%m-%d")
            except ValueError:
                errors.append(f"Row {row_idx}: invalid scheduled_date '{raw_date}' (expected YYYY-MM-DD)")
                continue

        priority = str(row_data.get("priority") or "NORMAL").strip().upper()
        if priority not in _PRIORITY_VALUES:
            priority = "NORMAL"

        def _parse_int(val: Any) -> int | None:
            if val is None or str(val).strip() == "":
                return None
            try:
                return int(float(val))
            except (ValueError, TypeError):
                return None

        try:
            order = Order(
                tenant_id=tenant_id,
                external_ref=str(row_data.get("external_ref") or "").strip() or None,
                delivery_address=str(row_data["delivery_address"]).strip(),
                delivery_latitude=_parse_float(row_data.get("delivery_latitude")),
                delivery_longitude=_parse_float(row_data.get("delivery_longitude")),
                scheduled_date=scheduled_date,
                time_window_start=_parse_time(row_data.get("time_window_start")),
                time_window_end=_parse_time(row_data.get("time_window_end")),
                weight_kg=_parse_float(row_data.get("weight_kg")),
                quantity_units=_parse_int(row_data.get("quantity_units")),
                value=_parse_float(row_data.get("value")),
                priority=priority,
                requires_refrigeration=_parse_bool(row_data.get("requires_refrigeration")),
                notes=str(row_data.get("notes") or "").strip() or None,
                status="PENDING",
            )
            db.add(order)
            created += 1
        except Exception as exc:
            errors.append(f"Row {row_idx}: {exc}")

    if created:
        db.commit()

    return {"created": created, "errors": len(errors), "error_details": errors}


def import_drivers_from_csv(db: Session, tenant_id: UUID, file_bytes: bytes) -> dict:
    reader = csv.DictReader(io.StringIO(file_bytes.decode("utf-8")))
    created = 0
    errors: list[str] = []

    for row_idx, row in enumerate(reader, start=2):
        name = str(row.get("full_name") or "").strip()
        if not name:
            errors.append(f"Row {row_idx}: missing full_name")
            continue
        try:
            driver = Driver(
                tenant_id=tenant_id,
                full_name=name,
                phone=str(row.get("phone") or "").strip() or None,
                email=str(row.get("email") or "").strip() or None,
                license_number=str(row.get("license_number") or "").strip() or None,
                license_class=str(row.get("license_class") or "").strip() or None,
                default_shift_start=_parse_time(row.get("default_shift_start")),
                default_shift_end=_parse_time(row.get("default_shift_end")),
                is_active=_parse_bool(row.get("is_active", "Yes")),
            )
            db.add(driver)
            created += 1
        except Exception as exc:
            errors.append(f"Row {row_idx}: {exc}")

    if created:
        db.commit()

    return {"created": created, "errors": len(errors), "error_details": errors}


def import_vehicles_from_csv(db: Session, tenant_id: UUID, file_bytes: bytes) -> dict:
    reader = csv.DictReader(io.StringIO(file_bytes.decode("utf-8")))
    created = 0
    errors: list[str] = []

    for row_idx, row in enumerate(reader, start=2):
        reg = str(row.get("registration_number") or "").strip()
        if not reg:
            errors.append(f"Row {row_idx}: missing registration_number")
            continue
        try:
            vehicle = Vehicle(
                tenant_id=tenant_id,
                registration_number=reg,
                vehicle_type=str(row.get("vehicle_type") or "VAN").strip().upper() or "VAN",
                capacity_kg=_parse_float(row.get("capacity_kg")),
                capacity_units=int(float(row.get("capacity_units"))) if row.get("capacity_units") else None,
                is_refrigerated=_parse_bool(row.get("is_refrigerated", "No")),
                is_active=_parse_bool(row.get("is_active", "Yes")),
            )
            db.add(vehicle)
            created += 1
        except Exception as exc:
            errors.append(f"Row {row_idx}: {exc}")

    if created:
        db.commit()

    return {"created": created, "errors": len(errors), "error_details": errors}


def import_depots_from_csv(db: Session, tenant_id: UUID, file_bytes: bytes) -> dict:
    reader = csv.DictReader(io.StringIO(file_bytes.decode("utf-8")))
    created = 0
    errors: list[str] = []

    for row_idx, row in enumerate(reader, start=2):
        name = str(row.get("name") or "").strip()
        if not name:
            errors.append(f"Row {row_idx}: missing name")
            continue
        try:
            depot = Depot(
                tenant_id=tenant_id,
                name=name,
                address=str(row.get("address") or "").strip() or None,
                city=str(row.get("city") or "").strip() or None,
                state=str(row.get("state") or "").strip() or None,
                country=str(row.get("country") or "India").strip() or "India",
                pincode=str(row.get("pincode") or "").strip() or None,
                latitude=_parse_float(row.get("latitude")),
                longitude=_parse_float(row.get("longitude")),
                is_active=_parse_bool(row.get("is_active", "Yes")),
            )
            db.add(depot)
            created += 1
        except Exception as exc:
            errors.append(f"Row {row_idx}: {exc}")

    if created:
        db.commit()

    return {"created": created, "errors": len(errors), "error_details": errors}
