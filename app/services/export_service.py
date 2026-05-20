import csv
import io
from datetime import date
from typing import Optional
from uuid import UUID
from sqlalchemy.orm import Session
from sqlalchemy import select, func
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.models.driver import Driver
from app.models.vehicle import Vehicle
from app.models.depot import Depot
from app.models.order import Order
from app.models.route_plan import RoutePlan, Route, RouteStop


_HEADER_FILL = PatternFill("solid", fgColor="1E40AF")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _style_header(ws, headers: list[str]):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = max(len(h) + 4, 14)


def orders_to_excel(
    db: Session,
    tenant_id: UUID,
    plan_date: Optional[date] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
) -> bytes:
    from datetime import datetime, timedelta
    stmt = select(Order).where(Order.tenant_id == tenant_id)
    if plan_date:
        stmt = stmt.where(func.date(Order.scheduled_date) == plan_date)
    elif date_from or date_to:
        if date_from:
            stmt = stmt.where(Order.scheduled_date >= datetime.combine(date_from, datetime.min.time()))
        if date_to:
            stmt = stmt.where(Order.scheduled_date < datetime.combine(date_to, datetime.min.time()) + timedelta(days=1))
    stmt = stmt.order_by(Order.scheduled_date)
    orders = db.execute(stmt).scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    headers = [
        "External Ref", "Delivery Address", "Latitude", "Longitude",
        "Weight (kg)", "Quantity", "Value",
        "Time Window Start", "Time Window End", "Scheduled Date",
        "Status", "Priority", "Requires Refrigeration", "Notes",
    ]
    _style_header(ws, headers)

    for order in orders:
        ws.append([
            order.external_ref or "",
            order.delivery_address,
            order.delivery_latitude,
            order.delivery_longitude,
            order.weight_kg,
            order.quantity_units,
            float(order.value) if order.value is not None else "",
            str(order.time_window_start) if order.time_window_start else "",
            str(order.time_window_end) if order.time_window_end else "",
            order.scheduled_date.strftime("%Y-%m-%d") if order.scheduled_date else "",
            order.status,
            order.priority,
            "Yes" if order.requires_refrigeration else "No",
            order.notes or "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def plan_to_excel(db: Session, tenant_id: UUID, plan_id: UUID) -> bytes | None:
    plan = db.execute(
        select(RoutePlan).where(
            RoutePlan.id == plan_id,
            RoutePlan.tenant_id == tenant_id,
        )
    ).scalar_one_or_none()
    if not plan:
        return None

    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    for route in plan.routes:
        driver_name = route.driver.name if route.driver else "Unassigned"
        ws = wb.create_sheet(title=driver_name[:31])  # Excel sheet name max 31 chars

        headers = ["Stop #", "External Ref", "Delivery Address", "Status", "Est. Arrival", "Weight (kg)", "Priority"]
        _style_header(ws, headers)

        for stop in route.stops:
            order = stop.order
            ws.append([
                stop.sequence,
                order.external_ref if order else "",
                order.delivery_address if order else "",
                stop.status,
                stop.estimated_arrival or "",
                order.weight_kg if order else "",
                order.priority if order else "",
            ])

    if not wb.sheetnames:
        ws = wb.create_sheet("No Routes")
        ws.append(["No routes in this plan"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def drivers_to_csv(db: Session, tenant_id: UUID) -> bytes:
    drivers = db.execute(
        select(Driver).where(Driver.tenant_id == tenant_id)
    ).scalars().all()

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        "full_name", "phone", "email", "license_number", "license_class",
        "default_shift_start", "default_shift_end", "is_active",
    ])
    for d in drivers:
        writer.writerow([
            d.full_name,
            d.phone or "",
            d.email or "",
            d.license_number or "",
            d.license_class or "",
            str(d.default_shift_start) if d.default_shift_start else "",
            str(d.default_shift_end) if d.default_shift_end else "",
            "Yes" if d.is_active else "No",
        ])
    return buf.getvalue().encode("utf-8")


def vehicles_to_csv(db: Session, tenant_id: UUID) -> bytes:
    vehicles = db.execute(
        select(Vehicle).where(Vehicle.tenant_id == tenant_id)
    ).scalars().all()

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        "registration_number", "vehicle_type", "capacity_kg", "capacity_units",
        "is_refrigerated", "is_active",
    ])
    for v in vehicles:
        writer.writerow([
            v.registration_number,
            v.vehicle_type,
            v.capacity_kg if v.capacity_kg is not None else "",
            v.capacity_units if v.capacity_units is not None else "",
            "Yes" if v.is_refrigerated else "No",
            "Yes" if v.is_active else "No",
        ])
    return buf.getvalue().encode("utf-8")


def depots_to_csv(db: Session, tenant_id: UUID) -> bytes:
    depots = db.execute(
        select(Depot).where(Depot.tenant_id == tenant_id)
    ).scalars().all()

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        "name", "address", "city", "state", "country", "pincode",
        "latitude", "longitude", "is_active",
    ])
    for d in depots:
        writer.writerow([
            d.name,
            d.address or "",
            d.city or "",
            d.state or "",
            d.country or "",
            d.pincode or "",
            d.latitude if d.latitude is not None else "",
            d.longitude if d.longitude is not None else "",
            "Yes" if d.is_active else "No",
        ])
    return buf.getvalue().encode("utf-8")


def orders_import_template() -> bytes:
    """Return a blank orders import template with headers and two sample rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders Import"

    headers = [
        "external_ref", "delivery_address", "delivery_latitude", "delivery_longitude",
        "scheduled_date", "time_window_start", "time_window_end",
        "weight_kg", "quantity_units", "value",
        "priority", "requires_refrigeration", "notes",
    ]
    _style_header(ws, headers)

    ws.append([
        "ORD-001", "123 MG Road, Bangalore", 12.9716, 77.5946,
        "2026-05-20", "09:00", "17:00", 5.0, 2, 1500.00,
        "NORMAL", "No", "Leave at door",
    ])
    ws.append([
        "ORD-002", "456 Indiranagar, Bangalore", 12.9784, 77.6408,
        "2026-05-20", "10:00", "14:00", 2.5, 1, 750.00,
        "HIGH", "No", "Call on arrival",
    ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
